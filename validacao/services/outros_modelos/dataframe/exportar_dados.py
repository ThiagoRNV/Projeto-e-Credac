from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from datetime import datetime
import pandas as pd
from django.db import connection
import logging

logger = logging.getLogger(__name__)


class ExportarServices:
    def __init__(self, empresa_id, data_sped) -> None:
        self.empresa_id = empresa_id
        self.data_sped = data_sped

    def exportar_service(self):
        if not self.empresa_id:
            logger.error('Erro no parametro empresa_id, valor como None')
            return {'empresa_id_error': True}

        if not self.data_sped:
            logger.error('Erro na data, data não recebida como parametro')
            return {'dt_sped_error': True}

        try:
            data_sped_obj = datetime.strptime(self.data_sped, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            logger.error('Data SPED inválida. Deve ser no formato YYYY-MM-DD.')
            return {'formato_date': True}

        query_sped = (
            "SELECT "
            "  CASE d.ind_oper WHEN '0' THEN 'Entrada' WHEN '1' THEN 'Saida' ELSE '---' END AS tipo, "
            "  p.cnpj_cpf AS cnpj_cpf, "
            "  p.nome AS nome, "
            "  d.num_doc AS num_doc, "
            "  COALESCE(d.chv_cte, '---') AS chv_cte, "
            "  COALESCE(d.ser, '---') AS ser, "
            "  d.dt_doc AS dt_doc, "
            "  COALESCE(a.cfop, '---') AS cfop, "
            "  COALESCE(a.cst_icms, '---') AS cst_icms, "
            "  COALESCE(a.aliq_icms, 0) AS aliq_icms, "
            "  COALESCE(a.vl_opr, 0) AS vl_opr, "
            "  COALESCE(a.vl_bc_icms, 0) AS vl_bc_icms, "
            "  COALESCE(a.vl_icms, 0) AS vl_icms, "
            "  COALESCE(a.vl_red_bc, 0) AS vl_red_bc, "
            "  COALESCE(d.vl_doc, 0) AS vl_doc, "
            "  COALESCE(d.vl_serv, 0) AS vl_serv, "
            "  COALESCE(a.cod_obs, '---') AS cod_obs "
            "FROM validacao_registrotransported100 AS d "
            "LEFT JOIN validacao_participantes AS p "
            "  ON d.cod_part_id = p.id "
            "INNER JOIN validacao_registrotransported190 AS a "
            "  ON a.registro_d100_id = d.id "
            "WHERE d.empresa_id = %s "
            "  AND d.data_inicio_sped = %s "
            "ORDER BY p.cod_part, d.id, a.id "
        )

        query_energia = (
            "SELECT "
            "  CASE c.ind_oper WHEN '0' THEN 'Entrada' WHEN '1' THEN 'Saida' ELSE '---' END AS tipo, "
            "  p.cnpj_cpf AS cnpj_cpf, "
            "  p.nome AS nome, "
            "  c.num_doc AS num_doc, "
            "  COALESCE(c.chv_doce, '---') AS chv_doce, "
            "  COALESCE(c.ser, '---') AS ser, "
            "  c.dt_doc AS dt_doc, "
            "  COALESCE(a.cfop, '---') AS cfop, "
            "  COALESCE(a.cst_icms, '---') AS cst_icms, "
            "  COALESCE(a.aliq_icms, 0) AS aliq_icms, "
            "  COALESCE(a.vl_opr, 0) AS vl_opr, "
            "  COALESCE(a.vl_bc_icms, 0) AS vl_bc_icms, "
            "  COALESCE(a.vl_icms, 0) AS vl_icms, "
            "  COALESCE(a.vl_bc_icms_st, 0) AS vl_bc_icms_st, "
            "  COALESCE(a.vl_icms_st, 0) AS vl_icms_st, "
            "  COALESCE(a.vl_red_bc, 0) AS vl_red_bc, "
            "  COALESCE(c.vl_doc, 0) AS vl_doc, "
            "  COALESCE(c.vl_forn, 0) AS vl_forn, "
            "  COALESCE(a.cod_obs, '---') AS cod_obs "
            "FROM validacao_registroenergiac500 AS c "
            "LEFT JOIN validacao_participantes AS p "
            "  ON c.cod_part_id = p.id "
            "INNER JOIN validacao_registroenergiac590 AS a "
            "  ON a.registro_c500_id = c.id "
            "WHERE c.empresa_id = %s "
            "  AND c.data_inicio_sped = %s "
            "ORDER BY p.cod_part, c.id, a.id "
        )

        query_comunicacao = (
            "SELECT "
            "  CASE d.ind_oper WHEN '0' THEN 'Entrada' WHEN '1' THEN 'Saida' ELSE '---' END AS tipo, "
            "  p.cnpj_cpf AS cnpj_cpf, "
            "  p.nome AS nome, "
            "  d.num_doc AS num_doc, "
            "  COALESCE(d.ser, '---') AS ser, "
            "  d.dt_doc AS dt_doc, "
            "  COALESCE(a.cfop, '---') AS cfop, "
            "  COALESCE(a.cst_icms, '---') AS cst_icms, "
            "  COALESCE(a.aliq_icms, 0) AS aliq_icms, "
            "  COALESCE(a.vl_opr, 0) AS vl_opr, "
            "  COALESCE(a.vl_bc_icms, 0) AS vl_bc_icms, "
            "  COALESCE(a.vl_icms, 0) AS vl_icms, "
            "  COALESCE(a.vl_bc_icms_st, 0) AS vl_bc_icms_st, "
            "  COALESCE(a.vl_icms_st, 0) AS vl_icms_st, "
            "  COALESCE(a.vl_red_bc, 0) AS vl_red_bc, "
            "  COALESCE(d.vl_doc, 0) AS vl_doc, "
            "  COALESCE(d.vl_serv, 0) AS vl_serv, "
            "  COALESCE(a.cod_obs, '---') AS cod_obs "
            "FROM validacao_registrocomunicacaod500 AS d "
            "LEFT JOIN validacao_participantes AS p "
            "  ON d.cod_part_id = p.id "
            "INNER JOIN validacao_registrocomunicacaod590 AS a "
            "  ON a.registro_d500_id = d.id "
            "WHERE d.empresa_id = %s "
            "  AND d.data_inicio_sped = %s "
            "ORDER BY p.cod_part, d.id, a.id "
        )

        with connection.cursor() as cursor:
            cursor.execute(query_sped, [self.empresa_id, data_sped_obj])
            rows = cursor.fetchall()
            cursor.execute(query_energia, [self.empresa_id, data_sped_obj])
            rows_energia = cursor.fetchall()
            cursor.execute(query_comunicacao, [self.empresa_id, data_sped_obj])
            rows_comunicacao = cursor.fetchall()

        colunas = [
            'REG',
            'CNPJ/CPF',
            'Nome',
            'Número Doc',
            'Chave CT-e',
            'Série',
            'Data Emissão',
            'CFOP',
            'CST',
            'Alíquota ICMS',
            'Valor Operação',
            'Base ICMS',
            'Valor ICMS',
            'Redução BC / ICMS-ST',
            'Valor Documento',
            'Valor Serviço',
            'Cód. Observação',
        ]

        colunas_energia = [
            'REG',
            'CNPJ/CPF',
            'Nome',
            'Número Doc',
            'Chave DOC-e',
            'Série',
            'Data Emissão',
            'CFOP',
            'CST',
            'Alíquota ICMS',
            'Valor Operação',
            'Base ICMS',
            'Valor ICMS',
            'Base ICMS-ST',
            'Valor ICMS-ST',
            'Redução BC',
            'Valor Documento',
            'Valor Fornecido',
            'Cód. Observação',
        ]

        colunas_comunicacao = [
            'REG',
            'CNPJ/CPF',
            'Nome',
            'Número Doc',
            'Série',
            'Data Emissão',
            'CFOP',
            'CST',
            'Alíquota ICMS',
            'Valor Operação',
            'Base ICMS',
            'Valor ICMS',
            'Base ICMS-ST',
            'Valor ICMS-ST',
            'Redução BC',
            'Valor Documento',
            'Valor Serviço',
            'Cód. Observação',
        ]

        def preencher_planilha(ws, colunas_ws, rows_ws):
            ws.append(colunas_ws)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for row in rows_ws:
                ws.append(row)

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2

        wb = Workbook()
        ws = wb.active
        ws.title = "Transportes"
        preencher_planilha(ws, colunas, rows)

        ws_energia = wb.create_sheet(title="Energia")
        preencher_planilha(ws_energia, colunas_energia, rows_energia)

        ws_comunicacao = wb.create_sheet(title="Comunicação")
        preencher_planilha(ws_comunicacao, colunas_comunicacao, rows_comunicacao)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        data_formatada = data_sped_obj.strftime("%d-%m-%Y")
        response['Content-Disposition'] = f'attachment; filename="relatorio_servicos_{data_formatada}.xlsx"'

        wb.save(response)
        return response
