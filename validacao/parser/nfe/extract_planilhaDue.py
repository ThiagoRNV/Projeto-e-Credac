from openpyxl import load_workbook

class ExtractPlanilhaDue:

    def __init__(self, planilha_due):
        self.planilha_due = planilha_due
        self.workbook = None

    def load_planilha_due(self):
        try:
            self.workbook = load_workbook(self.planilha_due)
            return True
        except Exception as e:
            print(f"Erro ao carregar planilha de due: {e}")
            return False

    def extract_values_planilha(self):
        if self.workbook is None:
            return []

        ws = self.workbook.worksheets[0]
        resultado = []

        # Layout da planilha (A ou B)


        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True):
            if not any(row):
                continue

            row_list = list(row) if row else []
            
            if row_list[0] is not None:
                data = row_list[0]
            else:
                data = row_list[1]

            list_mes = {
                 1 : 'Janeiro',
                 2 : 'Fevereiro',
                 3 : 'Março',
                 4 : 'Abril',
                 5 : 'Maio',
                 6 : 'Junho',
                 7 : 'Julho',
                 8 : 'Agosto',
                 9 : 'Setembro',
                 10 : 'Outubro',
                 11 : 'Novembro',
                 12 : 'Dezembro'
            } 
            mes = data.month
            mes_nome = list_mes.get(mes)
            try:
                if row_list[0] is not None:
                    if len(row_list) < 7:
                        continue
                    resultado.append({
                        'mes': mes_nome,
                        'nota': row_list[1] if len(row_list) > 1 else None,
                        'numero_due': row_list[4] if len(row_list) > 4 else None,
                    })
                else:
                    if len(row_list) < 6:
                        continue
                    resultado.append({
                        'mes': mes_nome,
                        'nota': row_list[2] if len(row_list) > 2 else None,
                        'numero_due': row_list[5] if len(row_list) > 5 else None,
                    })
            except (IndexError, TypeError) as e:
                # Pula linhas com problemas de índice
                print(f"Erro ao processar linha da planilha DUE: {e}")
                continue

        return {'planilha_due': resultado}
