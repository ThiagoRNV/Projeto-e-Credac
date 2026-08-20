from openpyxl import load_workbook

class ExtractPlanilhaCusto:

    def __init__(self, planilha_custo):
        self.planilha_custo = planilha_custo
        self.workbook = None
        

    def load_planilha_custo(self):
        try:
            self.workbook = load_workbook(self.planilha_custo)
            return True
        except Exception as e:
            print(f"Planilha inválida {e}")
            return False

    def extract_values_planilha(self):
        if self.workbook is None:
            return []
        ws = self.workbook.active
        self.planilha_custo = []
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True):
            self.planilha_custo.append({
                'categoria': row[1], 
                'centro_custo': row[2], 
                'descricao': row[3],
                'documento_fiscal': row[4],
                'fornecedor': row[5],
                'conta_contabil': row[6],
                'valor_total': row[7],
                'percentual_aplicado': row[8],
                'valor_alocado': row[9],
                'icms_passivel_credito': row[10],
            })
        return {'planilha_custo': self.planilha_custo}